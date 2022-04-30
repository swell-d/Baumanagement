from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from xlsx2html import xlsx2html

from Baumanagement.models.models_contracts import Bill, Contract


def tmp(request, id):
    bill = get_object_or_404(Bill, id=id)

    template = r"Baumanagement/print_forms/bill.xlsx"
    Path("tmp").mkdir(parents=True, exist_ok=True)
    newfilename = f'tmp/bill-{bill.id}.xlsx'

    seller = bill.contract.project.company if bill.contract.type == Contract.SELL else bill.contract.company
    company = bill.contract.company if bill.contract.type == Contract.SELL else bill.contract.project.company
    replace_dict = {
        'seller.name': seller.name,
        'seller.address': seller.address,
        'seller.city': seller.city,

        'company.name': company.name,
        'company.address': company.address,
        'company.city': company.city,
        'company.land': company.land,

        'bill.name': bill.name,
        'bill.date': bill.date.strftime("%d.%m.%Y") if bill.date else '',
        'company.id': company.id,
        'company.contact': seller.contacts.first(),

        'pos': '1',
        'item': 'Musterposition',
        'count': '1',
        'item.amount_netto_positiv': bill.amount_netto_positiv,

        'bill.amount_netto_positiv': bill.amount_netto_positiv,
        'vat': bill.vat,
        'bill.amount_brutto_positiv': bill.amount_brutto_positiv,

        'seller.ceo': seller.ceo,
    }

    def fill_fields(cell):
        for key, value in replace_dict.items():
            if key:
                cell = cell.replace('{' + key + '}', str(value or ''))
        return cell.strip()

    wb = openpyxl.load_workbook(template)
    sheet = wb.active
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if not sheet.cell(row=row, column=col).value:
                continue
            sheet.cell(row=row, column=col).value = fill_fields(sheet.cell(row=row, column=col).value)

    wb.save(newfilename)

    xlsx2html(newfilename, f'tmp/bill-{bill.id}.html')
    with open(f'tmp/bill-{bill.id}.html', 'r') as file:
        html = file.read()

    soup = BeautifulSoup(html, 'html.parser')
    soup.body['style'] = 'background-color: gray; font-family: sans-serif;'
    soup.body['onload'] = 'window.print();'
    soup.body.table['style'] = 'background-color: white;'

    return HttpResponse(soup.prettify())
