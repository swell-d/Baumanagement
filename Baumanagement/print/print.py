import openpyxl
from xlsx2html import xlsx2html

template = r"bill.xlsx"
newfilename = 'test.xlsx'

wb = openpyxl.load_workbook(template)
sheet = wb.active
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value:
            sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value.replace('{name}', 'dmitry')

wb.save(newfilename)

xlsx2html(newfilename, 'test.html')
