from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup
from django.contrib.auth.decorators import login_required
from django.http import FileResponse
from django.shortcuts import render, get_object_or_404
from xlsx2html import xlsx2html

from Baumanagement.models.models_bills import Bill
from contracts.models import Contract


def generate_excel(id):
    bill = get_object_or_404(Bill, id=id)

    template = r"print_forms/templates/bill.xlsx"
    Path("tmp").mkdir(parents=True, exist_ok=True)
    newfilename = f'tmp/bill-{id}.xlsx'

    seller = bill.contract.project.company if bill.contract.type == Contract.SELL else bill.contract.company
    company = bill.contract.company if bill.contract.type == Contract.BUY else bill.contract.project.company

    replace_dict = {
        'seller.name': seller.name,
        'seller.address': seller.address,
        'seller.city': seller.city,

        'company.name': company.name,
        'company.address': company.address,
        'company.city': company.city,
        'company.land': company.land,

        'bill.name': bill.name,
        'bill.date': bill.date.strftime("%d.%m.%Y") if bill.date else '',
        'company.id': company.id,

        'item.pos': '1',
        'item.name': 'Musterposition',
        'item.amount_netto_positiv': bill.amount_netto_positiv,
        'count': '1',
        'item.sum': bill.amount_netto_positiv,

        'bill.amount_netto_positiv': bill.amount_netto_positiv,
        'vat': bill.vat,
        'bill.amount_brutto_positiv': bill.amount_brutto_positiv,
        'bill.currency.symbol': bill.currency.symbol,

        'seller.phone': seller.phone,
        'seller.email': seller.email,

        'bill.account_to.IBAN': seller.accounts.first().IBAN,
        'bill.account_to.BIC': seller.accounts.first().BIC,
        'bill.account_to.bank': seller.accounts.first().bank,

        'seller.vat_number': seller.vat_number,
        'seller.ceo': seller.ceo,
    }

    def fill_fields(cell):
        for key, value in replace_dict.items():
            if key:
                cell = cell.replace('{' + key + '}', str(value or ''))
        return cell.strip()

    wb = openpyxl.load_workbook(template)
    sheet = wb.active
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if not sheet.cell(row=row, column=col).value:
                continue
            sheet.cell(row=row, column=col).value = fill_fields(sheet.cell(row=row, column=col).value)

    wb.save(newfilename)
    return newfilename


@login_required
def xlsx(request, id):
    newfilename = generate_excel(id)
    return FileResponse(open(newfilename, 'rb'))


@login_required
def html(request, id):
    newfilename = generate_excel(id)
    xlsx2html(newfilename, f'tmp/bill-{id}.html')
    with open(f'tmp/bill-{id}.html', 'r') as file:
        html = file.read()

    soup = BeautifulSoup(html, 'html.parser')
    for i, each in enumerate(soup.table.colgroup.findAll('col')):
        each['style'] = f'width: {210 / 12}mm'
        if i > 11:
            each.extract()
    soup.body.table.findAll('tr')[-1].extract()
    soup.body.table.findAll('tr')[-1].extract()
    soup.body.table['style'] = 'width: 210mm; height: 297mm; border-collapse: collapse; background-color: white;'

    context = {
        'title': f'bill-{id}',
        'html': str(soup.body.table)
    }
    return render(request, 'print_forms/print_form.html', context)
